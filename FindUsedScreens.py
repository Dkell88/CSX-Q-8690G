from pathlib import Path
from openpyxl import load_workbook

# ==== USER SETTINGS (you can leave these as just filenames) ====
HMI_FILE_RAW = "HMI Tag Format Example.xlsx"
WW_FILE_RAW  = "ww_TAG_crossref Example.xlsx"  # note: trim spaces is handled below

HMI_SHEET = "Progress Tracker"
WW_SHEET  = "HMI SCREENS, ANIMATIONS"

HMI_COL_SCREEN = "A"   # source names (with header)
HMI_COL_STATUS = "C"   # write results here

# Matching behavior
CASE_INSENSITIVE = True
SUBSTRING_MATCH  = True   # True: substring match; False: exact match

# ==== helpers ====
def norm(s: str) -> str:
    s = (s or "").strip()
    return s.lower() if CASE_INSENSITIVE else s

def resolve_file(name_or_path: str) -> Path:
    """
    Try to resolve a file by:
      1) Direct path
      2) Current working directory
      3) Script directory
      4) Recursive search under CWD, then under script dir (limited depth)
    Raises FileNotFoundError with helpful hints if not found.
    """
    clean = name_or_path.strip()
    candidate = Path(clean)

    # 1) direct / absolute
    if candidate.is_file():
        return candidate.resolve()

    # common search roots
    cwd = Path.cwd()
    script_dir = Path(__file__).parent.resolve()

    # 2) cwd
    if (cwd / candidate).is_file():
        return (cwd / candidate).resolve()

    # 3) script dir
    if (script_dir / candidate).is_file():
        return (script_dir / candidate).resolve()

    # 4) quick recursive searches (limit to a sane size)
    #   Search in CWD first
    try:
        found = next(cwd.rglob(candidate.name))
        if found.is_file():
            return found.resolve()
    except StopIteration:
        pass
    #   Then search in script directory
    try:
        found = next(script_dir.rglob(candidate.name))
        if found.is_file():
            return found.resolve()
    except StopIteration:
        pass

    # Not found -> build a helpful message
    tips = [
        f"Looked for '{candidate.name}' in:",
        f"  - {cwd}",
        f"  - {script_dir}",
        "…and via a quick recursive search under those folders.",
        "Fixes:",
        "  • Use the full absolute path, e.g.: r'C:\\path\\to\\HMI Tag Format Example.xlsx'",
        "  • Or place the script next to the Excel files.",
        "  • Or run the script from the folder that contains the files.",
    ]
    raise FileNotFoundError("\n".join(tips))

def load_ww_values_excluding_col_a(path: Path, sheet_name: str):
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {path.name}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    vals = []
    max_col = ws.max_column or 1
    # Exclude column A => start at column 2
    for row in ws.iter_rows(min_row=1, min_col=2, max_col=max_col, values_only=True):
        for v in row:
            if v is None:
                continue
            s = str(v).strip()
            if s:
                vals.append(norm(s))
    wb.close()
    return vals

def any_match(needle: str, haystack_values: list[str]) -> bool:
    n = norm(needle)
    if not n:
        return False
    if SUBSTRING_MATCH:
        return any(n in h for h in haystack_values)
    else:
        # exact match path uses a set for speed
        global _exact_set_cache
        try:
            _exact_set_cache
        except NameError:
            _exact_set_cache = set(haystack_values)
        return n in _exact_set_cache

def main():
    # Resolve files (handles stray spaces on names)
    hmi_path = resolve_file(HMI_FILE_RAW)
    ww_path  = resolve_file(WW_FILE_RAW)

    print(f"HMI workbook: {hmi_path}")
    print(f"WW  workbook: {ww_path}")

    # Preload WW values excluding column A
    ww_values = load_ww_values_excluding_col_a(ww_path, WW_SHEET)
    print(f"Loaded {len(ww_values)} non-column-A cell values from '{WW_SHEET}'.")

    # Open HMI workbook for editing
    hmi_wb = load_workbook(hmi_path, data_only=False)
    if HMI_SHEET not in hmi_wb.sheetnames:
        hmi_wb.close()
        raise ValueError(f"Sheet '{HMI_SHEET}' not found in {hmi_path.name}. Available: {hmi_wb.sheetnames}")
    hmi_ws = hmi_wb[HMI_SHEET]

    # Make sure Column C has a header
    if not (hmi_ws[f"{HMI_COL_STATUS}1"].value or "").strip():
        hmi_ws[f"{HMI_COL_STATUS}1"].value = "Status"

    # Iterate screen names in column A (skip header at row 1)
    updated = 0
    row = 2
    max_row = hmi_ws.max_row or 2
    # We’ll go to max_row but also continue if we encounter non-empty later rows
    while row <= max_row or any((hmi_ws[f"{HMI_COL_SCREEN}{r}"].value is not None)
                                for r in range(max_row+1, max_row+51)):
        # extend scan window slightly in case data_only/max_row is stale
        if row > max_row:
            max_row += 1

        name_cell = hmi_ws[f"{HMI_COL_SCREEN}{row}"]
        if name_cell.value is None:
            row += 1
            continue

        screen_name = str(name_cell.value).strip()
        if screen_name:
            found = any_match(screen_name, ww_values)
            hmi_ws[f"{HMI_COL_STATUS}{row}"].value = "Used" if found else "Not - Called"
            updated += 1
        else:
            hmi_ws[f"{HMI_COL_STATUS}{row}"].value = None
        row += 1

    hmi_wb.save(hmi_path)
    hmi_wb.close()
    print(f"Updated {updated} rows in '{hmi_path.name}' → sheet '{HMI_SHEET}', column {HMI_COL_STATUS}.")

if __name__ == "__main__":
    main()
